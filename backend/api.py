#!/usr/bin/env python3
"""
FastAPI Backend for GST Automation MVP

Provides REST API endpoints for:
- Invoice upload and processing
- HSN code suggestions
- GSTR-3B generation
- Invoice listing
"""

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Body, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from pathlib import Path
from typing import List, Optional, Dict
import json
import shutil
import uuid
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.processor import InvoiceProcessor
from app.hsn_matcher import HSNMatcher
from app.auth import get_user_id
from app.db import db

# Initialize FastAPI
app = FastAPI(
    title="GST Automation API",
    description="Backend API for GST invoice processing and GSTR-3B generation",
    version="1.0.0"
)

# CORS - Allow frontend to access API
import os
frontend_url = os.getenv("FRONTEND_URL", "http://localhost:3000")
origins = [frontend_url, "http://localhost:3000", "http://localhost:3001"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize processor
processor = InvoiceProcessor()
hsn_matcher = HSNMatcher()

# Data directory
DATA_DIR = Path(__file__).parent / "data"
INVOICES_DIR = DATA_DIR / "invoices"
UPLOADS_DIR = Path(__file__).parent / "uploads"
ALLOWED_EXTENSIONS = {'.pdf', '.jpg', '.jpeg', '.png', '.docx', '.doc', '.xlsx', '.xls'}

# Ensure directories exist
INVOICES_DIR.mkdir(parents=True, exist_ok=True)
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)


def validate_file_extension(filename: str) -> None:
    """Validate uploaded invoice file extension."""
    if not filename:
        raise HTTPException(400, "No filename provided")

    file_ext = Path(filename).suffix.lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            400,
            f"Invalid file type. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
        )


def build_unique_upload_path(filename: str) -> Path:
    """Build unique upload path to avoid collisions between files with same names."""
    safe_filename = Path(filename).name
    file_ext = Path(safe_filename).suffix.lower()
    unique_name = f"{uuid.uuid4().hex}{file_ext}"
    return UPLOADS_DIR / unique_name


@app.get("/")
async def root():
    """Health check endpoint"""
    return {
        "status": "operational",
        "version": "1.0.0",
        "service": "GST Automation API"
    }


@app.get("/api/health")
async def health_check():
    """Detailed health check"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "components": {
            "ocr": "ready",
            "hsn_matcher": "ready",
            "gstr3b_generator": "ready"
        }
    }


@app.post("/api/upload-bulk")
async def upload_bulk_invoices(
    files: List[UploadFile] = File(...),
    invoice_type: str = Form(...),
    user_id: str = Depends(get_user_id)
):
    """
    Upload and process multiple invoices in parallel
    """
    import asyncio
    
    print(f"\n{'='*60}")
    print(f"Bulk Processing: {len(files)} files for user {user_id}")
    print(f"{'='*60}\n")

    # 1. Save all files first (sanitize filenames)
    saved_files = []
    for file in files:
        validate_file_extension(file.filename)
        file_path = build_unique_upload_path(file.filename)
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        saved_files.append((Path(file.filename).name, str(file_path)))

    # 2. Define async processing wrapper
    async def process_single(filename, path):
        try:
            # Run blocking code in thread pool
            result = await asyncio.to_thread(
                processor.process_invoice,
                path,
                invoice_type=invoice_type,
                auto_confirm=True
            )
            
            # Add user_id to result and save updated file
            if result.get('success'):
                result['user_id'] = user_id
                if 'invoice_id' in result:
                    invoice_path = INVOICES_DIR / f"{result['invoice_id']}.json"
                    if invoice_path.exists():
                        with open(invoice_path, 'w') as f:
                            json.dump(result, f, indent=2, default=str)
            
            return {
                "filename": filename,
                "success": result['success'],
                "invoice_id": result.get('invoice_id'),
                "error": result.get('error')
            }
        except Exception as e:
            print(f"❌ Error processing {filename}: {str(e)}")
            return {
                "filename": filename,
                "success": False,
                "error": str(e)
            }

    # 3. Run all in parallel
    tasks = [process_single(fname, fpath) for fname, fpath in saved_files]
    results = await asyncio.gather(*tasks)

    # 4. Aggregate results
    success_count = sum(1 for r in results if r['success'])
    failed_count = len(results) - success_count

    return {
        "success": True,
        "message": f"Processed {len(files)} files: {success_count} success, {failed_count} failed",
        "summary": {
            "total": len(files),
            "success": success_count,
            "failed": failed_count
        },
        "results": results
    }


@app.post("/api/upload-invoice")
async def upload_invoice(
    file: UploadFile = File(...),
    invoice_type: str = Form("purchase"),
    user_id: str = Depends(get_user_id)
):
    """
    Upload and process an invoice

    Args:
        file: Invoice PDF/image file
        invoice_type: 'sales' or 'purchase'
        user_id: Authenticated user ID

    Returns:
        Processed invoice data with HSN categorization
    """
    try:
        # Validate file type
        validate_file_extension(file.filename)

        # Save uploaded file (sanitize filename to prevent path traversal)
        upload_path = build_unique_upload_path(file.filename)
        with open(upload_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Process invoice
        result = processor.process_invoice(
            str(upload_path),
            invoice_type=invoice_type,
            auto_confirm=True  # Auto-confirm high confidence matches
        )

        if not result.get('success'):
            raise HTTPException(500, result.get('error', 'Processing failed'))
            
        # Add user_id to result and save updated file
        result['user_id'] = user_id
        
        # We need to update the saved JSON file with user_id
        if 'invoice_id' in result:
            invoice_path = INVOICES_DIR / f"{result['invoice_id']}.json"
            if invoice_path.exists():
                with open(invoice_path, 'w') as f:
                    json.dump(result, f, indent=2, default=str)

        return {
            "success": True,
            "message": "Invoice processed successfully",
            "data": result
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Server error: {str(e)}")


@app.get("/api/invoices")
async def list_invoices(user_id: str = Depends(get_user_id)):
    """Get list of invoices for the authenticated user"""
    try:
        invoices = []

        if INVOICES_DIR.exists():
            for invoice_file in INVOICES_DIR.glob("*.json"):
                try:
                    with open(invoice_file, 'r') as f:
                        invoice_data = json.load(f)
                        # Only include invoices belonging to the user
                        # For backward compatibility, if no user_id is set in invoice, show it (or hide it)
                        # For now, we'll hide invoices without matching user_id to ensure isolation
                        if invoice_data.get('user_id') == user_id:
                            invoices.append(invoice_data)
                except Exception as e:
                    print(f"Error reading {invoice_file}: {e}")

        # Sort by date (newest first)
        invoices.sort(
            key=lambda x: x.get('invoice_date', ''),
            reverse=True
        )

        return {
            "success": True,
            "count": len(invoices),
            "invoices": invoices
        }

    except Exception as e:
        raise HTTPException(500, f"Error fetching invoices: {str(e)}")


@app.get("/api/invoices/export/excel")
async def export_invoices_excel(user_id: str = Depends(get_user_id)):
    """Export invoices for the authenticated user to Excel file"""
    try:
        invoices = []

        if INVOICES_DIR.exists():
            for invoice_file in INVOICES_DIR.glob("*.json"):
                try:
                    with open(invoice_file, 'r') as f:
                        invoice_data = json.load(f)
                        if invoice_data.get('user_id') == user_id:
                            invoices.append(invoice_data)
                except Exception as e:
                    print(f"Error reading {invoice_file}: {e}")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = ["Date", "Invoice #", "Type", "Vendor GSTIN", "Total Amount", 
                   "CGST", "SGST", "IGST", "Total Tax", "GST Rate", "Reconciliation Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, inv in enumerate(invoices, 2):
            ws.cell(row=row_idx, column=1, value=inv.get('invoice_date', ''))
            ws.cell(row=row_idx, column=2, value=inv.get('invoice_number', ''))
            ws.cell(row=row_idx, column=3, value=inv.get('invoice_type', ''))
            ws.cell(row=row_idx, column=4, value=inv.get('vendor_gstin', ''))
            ws.cell(row=row_idx, column=5, value=inv.get('total_amount', 0))
            ws.cell(row=row_idx, column=6, value=inv.get('cgst_amount', 0))
            ws.cell(row=row_idx, column=7, value=inv.get('sgst_amount', 0))
            ws.cell(row=row_idx, column=8, value=inv.get('igst_amount', 0))
            
            total_tax = (inv.get('cgst_amount', 0) + inv.get('sgst_amount', 0) + 
                        inv.get('igst_amount', 0))
            ws.cell(row=row_idx, column=9, value=total_tax)
            ws.cell(row=row_idx, column=10, value=inv.get('gst_rate', ''))
            ws.cell(row=row_idx, column=11, value=inv.get('reconciliation_status', 'pending'))
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save to file
        export_path = DATA_DIR / "invoices_export.xlsx"
        wb.save(export_path)
        
        return FileResponse(
            path=str(export_path),
            filename=f"invoices_{datetime.now().strftime('%Y%m%d')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        raise HTTPException(500, f"Error exporting to Excel: {str(e)}")


@app.get("/api/invoice/{invoice_id}")
async def get_invoice(invoice_id: str, user_id: str = Depends(get_user_id)):
    """Get details of a specific invoice"""
    try:
        invoice_path = INVOICES_DIR / f"{invoice_id}.json"

        if not invoice_path.exists():
            raise HTTPException(404, "Invoice not found")

        with open(invoice_path, 'r') as f:
            invoice_data = json.load(f)
            
        # Verify ownership
        if invoice_data.get('user_id') != user_id:
            raise HTTPException(403, "Access denied")

        return {
            "success": True,
            "invoice": invoice_data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error fetching invoice: {str(e)}")


@app.post("/api/suggest-hsn")
async def suggest_hsn(item_description: str = Form(...)):
    """
    Get HSN code suggestions for an item description

    Args:
        item_description: Description of the item/service

    Returns:
        HSN code suggestions with confidence scores
    """
    try:
        suggestions = hsn_matcher.suggest_hsn(item_description, top_n=3)

        return {
            "success": True,
            "item_description": item_description,
            "suggestions": suggestions
        }

    except Exception as e:
        raise HTTPException(500, f"Error suggesting HSN: {str(e)}")


@app.patch("/api/invoice/{invoice_id}")
async def update_invoice(invoice_id: str, update_data: Dict = Body(...), user_id: str = Depends(get_user_id)):
    """
    Update invoice fields (e.g. reconciliation_status)
    """
    try:
        # Load invoice
        invoice_path = processor.invoices_dir / f"{invoice_id}.json"
        if not invoice_path.exists():
            raise HTTPException(404, "Invoice not found")

        with open(invoice_path, "r") as f:
            invoice = json.load(f)

        # Verify ownership before allowing updates
        if invoice.get('user_id') != user_id:
            raise HTTPException(403, "Access denied")

        # Update allowed fields
        allowed_fields = [
            'reconciliation_status',
            'invoice_number',
            'vendor_gstin',
            'invoice_date',
            'total_amount',
            'gst_rate'
        ]
        for key, value in update_data.items():
            if key in allowed_fields:
                invoice[key] = value

        # Keep tax fields in sync when amount/rate is edited.
        if 'total_amount' in update_data or 'gst_rate' in update_data:
            total_amount = float(invoice.get('total_amount') or 0)
            gst_rate = float(invoice.get('gst_rate') or 0)
            if total_amount > 0 and gst_rate > 0:
                total_tax = total_amount * (gst_rate / (100 + gst_rate))

                user_gstin = os.environ.get('USER_GSTIN', '')
                user_state_code = user_gstin[:2] if len(user_gstin) >= 2 else "32"

                vendor_gstin = invoice.get('vendor_gstin')
                vendor_state_code = vendor_gstin[:2] if vendor_gstin and len(vendor_gstin) >= 2 else None

                if vendor_state_code and vendor_state_code != user_state_code:
                    igst = round(total_tax, 2)
                    cgst = 0
                    sgst = 0
                else:
                    cgst = round(total_tax / 2, 2)
                    sgst = round(total_tax - cgst, 2)
                    igst = 0

                invoice['cgst'] = cgst
                invoice['sgst'] = sgst
                invoice['igst'] = igst
                invoice['cgst_amount'] = cgst
                invoice['sgst_amount'] = sgst
                invoice['igst_amount'] = igst

        # Save back
        with open(invoice_path, "w") as f:
            json.dump(invoice, f, indent=2)

        # Also update Supabase if connected
        if db.is_connected():
            # This is a simplified update, ideally we'd have a specific update method in db.py
            # For now, we'll just print a warning that DB update isn't fully implemented for PATCH
            print("⚠️  Note: Database update for PATCH not yet implemented")

        return {"success": True, "message": "Invoice updated", "data": invoice}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error updating invoice: {str(e)}")


@app.post("/api/generate-gstr3b")
async def generate_gstr3b(
    gstin: str = Form(...),
    month: int = Form(...),
    year: int = Form(...),
    user_id: str = Depends(get_user_id)
):
    """
    Generate GSTR-3B monthly return

    Args:
        gstin: Your GSTIN (15 characters)
        month: Month (1-12)
        year: Year (YYYY)

    Returns:
        Generated GSTR-3B data
    """
    try:
        # Validate inputs
        if len(gstin) != 15:
            raise HTTPException(400, "GSTIN must be 15 characters")

        if not (1 <= month <= 12):
            raise HTTPException(400, "Month must be between 1 and 12")

        if year < 2000 or year > 2100:
            raise HTTPException(400, "Invalid year")

        # Generate GSTR-3B
        result = processor.generate_gstr3b(month, year, gstin)

        if not result.get('success'):
            raise HTTPException(500, result.get('error', 'Generation failed'))

        return {
            "success": True,
            "message": "GSTR-3B generated successfully",
            "data": result['gstr3b'],
            "file_path": result.get('file_path')
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error generating GSTR-3B: {str(e)}")


@app.get("/api/stats")
async def get_stats(user_id: str = Depends(get_user_id)):
    """Get dashboard statistics for the authenticated user"""
    try:
        invoices = []

        if INVOICES_DIR.exists():
            for invoice_file in INVOICES_DIR.glob("*.json"):
                try:
                    with open(invoice_file, 'r') as f:
                        data = json.load(f)
                    if data.get('user_id') == user_id:
                        invoices.append(data)
                except Exception:
                    pass

        # Calculate stats
        total_invoices = len(invoices)
        total_amount = sum(inv.get('total_amount', 0) for inv in invoices)
        total_tax = sum(
            inv.get('cgst_amount', 0) +
            inv.get('sgst_amount', 0) +
            inv.get('igst_amount', 0)
            for inv in invoices
        )

        sales_invoices = [inv for inv in invoices if inv.get('invoice_type') == 'sales']
        purchase_invoices = [inv for inv in invoices if inv.get('invoice_type') == 'purchase']

        return {
            "success": True,
            "stats": {
                "total_invoices": total_invoices,
                "sales_count": len(sales_invoices),
                "purchase_count": len(purchase_invoices),
                "total_amount": round(total_amount, 2),
                "total_tax": round(total_tax, 2),
                "avg_processing_confidence": round(
                    sum(
                        inv.get('ocr_confidence', inv.get('confidence', 0))
                        for inv in invoices
                    ) / max(total_invoices, 1),
                    1
                )
            }
        }

    except Exception as e:
        raise HTTPException(500, f"Error fetching stats: {str(e)}")


@app.post("/api/reconcile-gstr2a")
async def reconcile_gstr2a(file: UploadFile = File(...), user_id: str = Depends(get_user_id)):
    """
    Reconcile uploaded invoices with GSTR-2A data
    
    This is a placeholder implementation that demonstrates the concept.
    In production, this would parse the actual GSTR-2A Excel/JSON file.
    """
    try:
        # Load invoices for this user
        our_invoices = []
        if INVOICES_DIR.exists():
            for invoice_file in INVOICES_DIR.glob("*.json"):
                try:
                    with open(invoice_file, 'r') as f:
                        invoice_data = json.load(f)
                        if (invoice_data.get('user_id') == user_id and
                                invoice_data.get('invoice_type') == 'purchase'):
                            our_invoices.append(invoice_data)
                except Exception as e:
                    print(f"Error reading {invoice_file}: {e}")
        
        # MVP: Show invoices that would need reconciliation
        # Real implementation would parse the GSTR-2A Excel/JSON file
        details = []
        for inv in our_invoices:
            details.append({
                'invoice_number': inv.get('invoice_number'),
                'vendor_gstin': inv.get('vendor_gstin'),
                'your_amount': inv.get('total_amount'),
                'portal_amount': None,
                'status': 'pending_reconciliation'
            })

        return {
            'success': True,
            'is_demo': True,
            'matched': 0,
            'missing_in_portal': 0,
            'amount_mismatch': 0,
            'total_invoices': len(our_invoices),
            'details': details,
            'message': 'DEMO MODE: GSTR-2A file parsing is not yet implemented. Showing your purchase invoices that would be reconciled.'
        }
    
    except Exception as e:
        raise HTTPException(500, f"Error reconciling: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)
